"""Ops-Hardening Phase 3. Daily trade-log Excel export — pure data-query +
file-write logic, no threading (`export_scheduler.py` is the thread that
calls this on a schedule, same split `market_data_scheduler.py`/
`market_hours.py` already use).

**Deliberately `openpyxl` only, not `pandas`** — see this module's own
dependency comment in `pyproject.toml`: pandas pulls in numpy, whose stub
files break `mypy app tests` under this project's `python_version = "3.11"`
target, which is exactly why pandas has been kept out of the main `.venv`
since before this phase. Nothing here needs DataFrame machinery — it's a
handful of columns per row, read via a plain SQLAlchemy query.

**One workbook per workspace** (`reports/trade_log_<workspace_id>.xlsx`),
not one shared file — matches `HealthCheckScheduler`'s own precedent of
never assuming single-tenant, and never mixes different workspaces'
financial data in one file.

**One tab per strategy** (`StrategyConfig.name`) — **2026-08-19, changed
from the original one-tab-per-(underlying, expiry_date)** — a strategy's
full trade history (across every underlying/expiry cycle it ever ran on,
and across both paper and live trades — see the new "Paper/Live" column)
in one place, since evaluating a strategy's own performance (and comparing
its paper vs. live behavior) is the actual question this report exists to
answer, not "what traded on a given expiry." `Underlying`/`Expiry` moved
from being implied by the sheet name to explicit row columns, since a
strategy's own tab can now span more than one of either.

**Idempotent appends** — every row carries its source `TradeOutcome.id` as
its last column; before appending, a tab's existing IDs are read and
already-present ones skipped. This is what makes it safe for the scheduler
(`export_scheduler.py`) to re-trigger a day it already exported (e.g. after
a restart) without producing duplicate rows — a re-run is a no-op, not a
retry hazard, matching this project's own stated invariant ("any new write
path that could plausibly run twice needs to reason about this explicitly").
"""

from __future__ import annotations

import csv
import logging
import uuid
from collections import defaultdict
from dataclasses import dataclass
from datetime import UTC, date, datetime, time, timedelta
from pathlib import Path

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.config.settings import BACKEND_ROOT_DIR
from app.core.clock import IST, now_ist, to_ist
from app.core.db.session import SessionFactory, session_scope
from app.domain.execution.models import Order, Position, PositionExitLeg, TradeOutcome
from app.domain.market.models import Instrument, OptionContract
from app.domain.strategy.models import StrategyConfig, StrategyRun, TradeIntent
from app.modules.strategy_engine.env_metrics import (
    compute_pcr,
    get_chain_data_as_of,
    get_contract_oi,
    get_vix_as_of,
)

logger = logging.getLogger("app.reporting.exporter")

REPORTS_DIR = BACKEND_ROOT_DIR / "reports"

# Multi-leg exit engine: "Leg"/"Leg Kind" were added after this file had
# already been exporting daily to a real, live `trade_log_<workspace_id>
# .xlsx` for months. A naive positional column append would break
# idempotency on that pre-existing file -- see `_sheet_headers`/
# `_existing_trade_ids`'s own docstrings for why every column is resolved
# by header *name* within each sheet's own row 1, not by a fixed global
# index, and why an already-existing sheet deliberately never gains these
# two new columns (it just keeps behaving exactly as it always has).
_HEADERS = [
    "Strategy",
    "Underlying",
    "Expiry",
    "Execution Mode",
    "Paper/Live",
    "Contract Symbol",
    "Option Type",
    "Strike",
    "Side",
    "Qty",
    "Entry Price",
    "Entry Time (IST)",
    "Exit Price",
    "Exit Time (IST)",
    "Exit Reason",
    "Realized PnL",
    "Slippage",
    "VIX (at entry)",
    "OI (at entry)",
    "PCR - OI (at entry)",
    "PCR - Volume (at entry)",
    "Leg",
    "Leg Kind",
    "Trade ID (internal)",
]
_TRADE_ID_HEADER = "Trade ID (internal)"


@dataclass(frozen=True)
class TradeLogRow:
    trade_outcome_id: uuid.UUID
    workspace_id: uuid.UUID
    strategy_name: str
    execution_mode: str
    # "paper" / "live" -- Order.mode on the position's own opening order,
    # not this system's SafeMode/session-level concept. Distinct from
    # execution_mode (auto/manual approval), which is orthogonal to this.
    trade_mode: str
    underlying_symbol: str
    contract_symbol: str
    option_type: str
    strike: float
    expiry_date: date
    side: str
    qty: int
    entry_price: float
    entry_time_utc: datetime
    exit_price: float
    exit_time_utc: datetime
    exit_reason: str
    realized_pnl: float
    slippage: float
    # VIX/PCR/OI environment metrics as of this trade's entry time, not
    # "current" -- see strategy_engine.env_metrics.get_env_metrics's own
    # docstring for the as_of_utc reconstruction this is built from. `None`
    # for any/all when nothing was known yet at that moment (e.g. a trade
    # that predates the VIX/PCR pipeline entirely). `oi` is the specific
    # traded contract's own raw open interest (env_metrics.get_contract_oi),
    # distinct from pcr_oi/pcr_vol, which are chain-wide aggregates.
    vix: float | None
    oi: int | None
    pcr_oi: float | None
    pcr_vol: float | None
    # Multi-leg exit engine: "<leg_index+1>/<total legs for this position>"
    # (e.g. "2/3") for a staged-exit leg; "1/1" for a legacy single-outcome
    # trade with no PositionExitLeg row at all.
    leg_label: str
    # PositionExitLeg.kind ("fixed_sl"/"sr_target"/"runner"/"custom"/
    # "single") for a staged leg; "—" for a legacy row with no
    # position_exit_leg_id.
    leg_kind: str


def _day_bounds_utc(target_date: date) -> tuple[datetime, datetime]:
    """IST-calendar-day boundaries, converted to UTC for the query filter —
    never a bare `utcnow()`/naive-date comparison, per this project's own
    timezone-strictness convention (`app.core.clock`'s own docstring).
    """
    start_ist = datetime.combine(target_date, time.min, tzinfo=IST)
    start_utc = start_ist.astimezone(UTC)
    return start_utc, start_utc + timedelta(days=1)


def fetch_completed_trades_for_day(db: Session, target_date: date) -> list[TradeLogRow]:
    """Reuses `TradeOutcome` — the same "completed trade" source of truth
    `reporting.service.build_daily_report`/`build_scorecard` already read —
    rather than re-deriving trade completion independently.
    """
    start_utc, end_utc = _day_bounds_utc(target_date)

    query_rows = (
        db.query(
            TradeOutcome,
            Position,
            OptionContract,
            Instrument,
            TradeIntent,
            StrategyRun,
            StrategyConfig,
            Order,
            PositionExitLeg,
        )
        .join(Position, TradeOutcome.position_id == Position.id)
        .join(OptionContract, Position.option_contract_id == OptionContract.id)
        .join(Instrument, OptionContract.instrument_id == Instrument.id)
        .join(TradeIntent, TradeOutcome.trade_intent_id == TradeIntent.id)
        .join(StrategyRun, TradeIntent.strategy_run_id == StrategyRun.id)
        .join(StrategyConfig, StrategyRun.strategy_config_id == StrategyConfig.id)
        .join(Order, Position.opening_order_id == Order.id)
        # Nullable -- None for a legacy single-outcome row with no
        # position_exit_leg_id (see TradeOutcome's own docstring).
        .outerjoin(PositionExitLeg, TradeOutcome.position_exit_leg_id == PositionExitLeg.id)
        .filter(TradeOutcome.closed_at >= start_utc, TradeOutcome.closed_at < end_utc)
        .order_by(TradeOutcome.closed_at)
        .all()
    )

    # One extra batched query (not per-row) for "how many legs does this
    # position have in total" -- same batch-by-position_id.in_(...) idiom
    # `api.v1.execution.list_positions` already uses for `legs_by_position`.
    position_ids = [position.id for _, position, *_rest in query_rows]
    leg_counts_by_position: dict[uuid.UUID, int] = {}
    if position_ids:
        for leg_position_id, leg_count in (
            db.query(PositionExitLeg.position_id, func.count())
            .filter(PositionExitLeg.position_id.in_(position_ids))
            .group_by(PositionExitLeg.position_id)
            .all()
        ):
            leg_counts_by_position[leg_position_id] = leg_count

    rows = []
    for (
        outcome,
        position,
        contract,
        instrument,
        intent,
        run,
        config,
        opening_order,
        exit_leg,
    ) in query_rows:
        # As of this trade's own entry time, not "current" -- see
        # get_chain_data_as_of/get_vix_as_of's own docstrings for the
        # as_of_utc reconstruction. Two extra queries per row (VIX tick +
        # option-chain snapshot); fine at this system's real trade volumes,
        # and this only ever runs once daily, off the hot path.
        as_of = position.opened_at
        vix = get_vix_as_of(db, as_of_utc=as_of)
        chain_data = get_chain_data_as_of(db, instrument.id, contract.expiry_date, as_of_utc=as_of)
        pcr_oi, pcr_vol = compute_pcr(chain_data) if chain_data is not None else (None, None)
        oi = get_contract_oi(chain_data, contract.symbol) if chain_data is not None else None
        if exit_leg is not None:
            leg_label = f"{exit_leg.leg_index + 1}/{leg_counts_by_position.get(position.id, 1)}"
            leg_kind = exit_leg.kind
        else:
            leg_label = "1/1"
            leg_kind = "—"
        rows.append(
            # `.value`, not a bare attribute -- these columns are all plain
            # String(N) (no native SQLAlchemy Enum type), so a value freshly
            # read back via this query's own SELECT comes back as a plain
            # `str`, not the declared StrEnum subtype (confirmed
            # empirically: a bare `.value` access AttributeErrors on the
            # read-back object even though the same field accepts an enum
            # member fine on construction). `str(...)` handles both shapes
            # uniformly rather than assuming one.
            TradeLogRow(
                trade_outcome_id=outcome.id,
                workspace_id=intent.workspace_id,
                strategy_name=config.name,
                execution_mode=str(run.execution_mode),
                trade_mode=str(opening_order.mode),
                underlying_symbol=instrument.symbol,
                contract_symbol=contract.symbol,
                option_type=str(contract.option_type),
                strike=float(contract.strike),
                expiry_date=contract.expiry_date,
                side=str(position.side),
                # `outcome.qty`, not `position.qty` -- the latter is
                # decremented toward 0 as each staged-exit leg closes (see
                # exit_legs.py), so it no longer reflects this specific
                # leg's own fill size once other legs have also closed.
                # `TradeOutcome.qty` is always the correct per-leg (or,
                # for a legacy trade, whole-position) quantity.
                qty=outcome.qty,
                entry_price=float(outcome.entry_price),
                entry_time_utc=position.opened_at,
                exit_price=float(outcome.exit_price),
                exit_time_utc=outcome.closed_at,
                exit_reason=str(outcome.exit_reason),
                realized_pnl=float(outcome.realized_pnl),
                slippage=float(outcome.slippage),
                vix=vix,
                oi=oi,
                pcr_oi=pcr_oi,
                pcr_vol=pcr_vol,
                leg_label=leg_label,
                leg_kind=leg_kind,
            )
        )
    return rows


def _sanitize_sheet_name(name: str) -> str:
    for ch in ":\\/?*[]":
        name = name.replace(ch, "-")
    return name[:31]


def _row_field_map(row: TradeLogRow) -> dict[str, object]:
    """Header label -> value, for every column this file has ever had or
    currently has. Resolving columns by *name* (see `_row_values_for_sheet`/
    `_existing_trade_ids` below) rather than a fixed position is what lets a
    header set grow over time (e.g. the "Leg"/"Leg Kind" addition) without
    corrupting a sheet that was created under an older, narrower `_HEADERS`.
    """
    return {
        "Strategy": row.strategy_name,
        "Underlying": row.underlying_symbol,
        "Expiry": row.expiry_date.isoformat(),
        "Execution Mode": row.execution_mode,
        "Paper/Live": row.trade_mode,
        "Contract Symbol": row.contract_symbol,
        "Option Type": row.option_type,
        "Strike": row.strike,
        "Side": row.side,
        "Qty": row.qty,
        "Entry Price": row.entry_price,
        "Entry Time (IST)": to_ist(row.entry_time_utc).replace(tzinfo=None),
        "Exit Price": row.exit_price,
        "Exit Time (IST)": to_ist(row.exit_time_utc).replace(tzinfo=None),
        "Exit Reason": row.exit_reason,
        "Realized PnL": row.realized_pnl,
        "Slippage": row.slippage,
        "VIX (at entry)": row.vix,
        "OI (at entry)": row.oi,
        "PCR - OI (at entry)": row.pcr_oi,
        "PCR - Volume (at entry)": row.pcr_vol,
        "Leg": row.leg_label,
        "Leg Kind": row.leg_kind,
        _TRADE_ID_HEADER: str(row.trade_outcome_id),
    }


def _row_values_for_sheet(row: TradeLogRow, headers: list[str]) -> list:
    """Builds the row to append in *this sheet's own* header order. A
    header in `headers` that this file no longer writes would map to
    `None` (doesn't happen in practice -- headers are only ever added, not
    renamed/removed); a header in `_row_field_map` that this particular
    sheet doesn't have yet (e.g. "Leg"/"Leg Kind" on a sheet created before
    they existed) is simply omitted for that sheet, which is the point --
    see `_HEADERS`'s own module-level comment.
    """
    field_map = _row_field_map(row)
    return [field_map.get(h) for h in headers]


def _sheet_headers(ws: Worksheet) -> list[str]:
    return [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]


def _existing_trade_ids(ws: Worksheet, headers: list[str]) -> set[str]:
    """Resolves the Trade-ID column by *this sheet's own* header row, not a
    fixed global index -- a sheet created under an older, narrower
    `_HEADERS` still has its real Trade ID data at its own original
    position, which a global `len(_HEADERS)`-based index would get wrong
    the moment new columns are added (see `_HEADERS`'s own comment). Falls
    back to an empty set (never crashes) if a sheet somehow has no such
    column at all -- shouldn't happen since every version of this file has
    always had it, but re-exporting into a genuinely foreign/hand-edited
    sheet should degrade to "nothing recognized as already exported"
    rather than raise.
    """
    if _TRADE_ID_HEADER not in headers:
        logger.warning(
            "Sheet has no %r column -- treating every row as not-yet-exported",
            _TRADE_ID_HEADER,
        )
        return set()
    column_index = headers.index(_TRADE_ID_HEADER)
    ids: set[str] = set()
    for row_cells in ws.iter_rows(min_row=2):
        cell = row_cells[column_index]
        if cell.value:
            ids.add(str(cell.value))
    return ids


def _get_or_create_sheet(wb: openpyxl.Workbook, sheet_name: str) -> Worksheet:
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    ws = wb.create_sheet(title=sheet_name)
    ws.append(_HEADERS)
    return ws


def _write_csv_fallback(
    workspace_id: uuid.UUID, rows: list[TradeLogRow], target_date: date
) -> Path:
    """The file-lock guard's actual safety net — same row shape as the
    Excel append, just to a fresh, never-locked file, so a `.xlsx` left open
    on the desktop at 15:35 never costs the day's records. No separate
    prepended Underlying/Expiry columns (unlike before 2026-08-19's
    per-strategy-tab change) -- both are now real `_HEADERS` columns
    themselves, since a flat CSV has no sheet to imply them from either way.
    """
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    stamp = now_ist().strftime("%H%M%S")
    path = REPORTS_DIR / f"trade_log_{workspace_id}_{target_date.isoformat()}_fallback_{stamp}.csv"
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(_HEADERS)
        for row in rows:
            # Always a brand-new file -- no pre-existing header to be
            # compatible with, so always the full current `_HEADERS`.
            writer.writerow(_row_values_for_sheet(row, _HEADERS))
    return path


def export_trade_log_for_workspace(
    workspace_id: uuid.UUID, rows: list[TradeLogRow], target_date: date
) -> Path | None:
    """`rows` must already be scoped to one workspace. Returns the path
    actually written to (the `.xlsx`, or a CSV fallback on a file-lock
    error), or `None` if there was nothing to export.
    """
    if not rows:
        logger.info("No trades to export for workspace %s on %s", workspace_id, target_date)
        return None

    REPORTS_DIR.mkdir(parents=True, exist_ok=True)
    path = REPORTS_DIR / f"trade_log_{workspace_id}.xlsx"

    is_new_workbook = not path.exists()
    wb = openpyxl.load_workbook(path) if not is_new_workbook else openpyxl.Workbook()
    if is_new_workbook:
        # A freshly-created Workbook() always starts with exactly one
        # default "Sheet" -- safe to drop immediately since a real sheet is
        # about to be created below (rows is non-empty, checked above).
        del wb["Sheet"]

    by_sheet: dict[str, list[TradeLogRow]] = defaultdict(list)
    for row in rows:
        by_sheet[_sanitize_sheet_name(row.strategy_name)].append(row)

    appended = 0
    for sheet_name, sheet_rows in by_sheet.items():
        ws = _get_or_create_sheet(wb, sheet_name)
        headers = _sheet_headers(ws)
        existing_ids = _existing_trade_ids(ws, headers)
        for row in sheet_rows:
            if str(row.trade_outcome_id) in existing_ids:
                continue
            ws.append(_row_values_for_sheet(row, headers))
            appended += 1

    if appended == 0:
        logger.info(
            "All %d trade(s) for workspace %s on %s were already exported — nothing new to append",
            len(rows),
            workspace_id,
            target_date,
        )
        return path

    try:
        wb.save(path)
    except PermissionError:
        fallback_path = _write_csv_fallback(workspace_id, rows, target_date)
        logger.warning(
            "Could not save %s (likely open in another program) — wrote a CSV fallback "
            "to %s instead so today's %d trade(s) aren't lost",
            path,
            fallback_path,
            len(rows),
        )
        return fallback_path

    logger.info(
        "Exported %d new trade row(s) for workspace %s to %s", appended, workspace_id, path
    )
    return path


def export_completed_trades_for_day(
    target_date: date | None = None,
    *,
    session_factory: SessionFactory = session_scope,
) -> None:
    """Entry point `export_scheduler.py`'s daily trigger calls. `target_date`
    defaults to `now_ist().date()` — always IST, never a bare `utcnow()`
    date, per this project's own timezone-strictness convention.
    """
    resolved_date = target_date if target_date is not None else now_ist().date()

    with session_factory() as db:
        rows = fetch_completed_trades_for_day(db, resolved_date)

    if not rows:
        logger.info("No trades to export for %s", resolved_date)
        return

    by_workspace: dict[uuid.UUID, list[TradeLogRow]] = defaultdict(list)
    for row in rows:
        by_workspace[row.workspace_id].append(row)

    for workspace_id, workspace_rows in by_workspace.items():
        export_trade_log_for_workspace(workspace_id, workspace_rows, resolved_date)
