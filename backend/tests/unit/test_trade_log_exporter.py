"""Ops-Hardening Phase 3: app.modules.reporting.exporter -- pure day-boundary/
sheet-naming logic plus the Excel-write path (sheet routing, idempotent
append, file-lock fallback), driven with synthetic `TradeLogRow`s against a
`tmp_path`-backed `REPORTS_DIR` rather than the DB -- `fetch_completed_trades_
for_day`'s own DB-query correctness is covered separately in
tests/integration/test_trade_log_export_query.py.

**2026-08-19**: sheets route by strategy name now, not (underlying,
expiry_date) -- see exporter.py's own module docstring for why.
"""

from __future__ import annotations

import csv
import uuid
from datetime import UTC, date, datetime

import openpyxl
import pytest

from app.modules.reporting import exporter
from app.modules.reporting.exporter import (
    _HEADERS,
    TradeLogRow,
    _day_bounds_utc,
    _sanitize_sheet_name,
    export_trade_log_for_workspace,
)


@pytest.fixture(autouse=True)
def _reports_dir(tmp_path, monkeypatch):
    monkeypatch.setattr(exporter, "REPORTS_DIR", tmp_path)
    return tmp_path


def _row(
    *,
    workspace_id: uuid.UUID | None = None,
    underlying: str = "NIFTY",
    expiry: date = date(2026, 8, 18),
    strategy: str = "orb",
    trade_mode: str = "paper",
    trade_id: uuid.UUID | None = None,
    pnl: float = 500.0,
    vix: float | None = 13.5,
    oi: int | None = 125000,
    pcr_oi: float | None = 0.9,
    pcr_vol: float | None = 1.1,
    leg_label: str = "1/1",
    leg_kind: str = "—",
) -> TradeLogRow:
    return TradeLogRow(
        trade_outcome_id=trade_id or uuid.uuid4(),
        workspace_id=workspace_id or uuid.uuid4(),
        strategy_name=strategy,
        execution_mode="auto",
        trade_mode=trade_mode,
        underlying_symbol=underlying,
        contract_symbol=f"{underlying}{expiry.strftime('%d%b%y').upper()}C24000",
        option_type="CE",
        strike=24000.0,
        expiry_date=expiry,
        side="buy",
        qty=25,
        entry_price=80.0,
        entry_time_utc=datetime(2026, 8, 18, 5, 0, tzinfo=UTC),
        exit_price=100.0,
        exit_time_utc=datetime(2026, 8, 18, 6, 0, tzinfo=UTC),
        exit_reason="target",
        realized_pnl=pnl,
        slippage=1.5,
        vix=vix,
        oi=oi,
        pcr_oi=pcr_oi,
        pcr_vol=pcr_vol,
        leg_label=leg_label,
        leg_kind=leg_kind,
    )


# -- _day_bounds_utc ----------------------------------------------------------


def test_day_bounds_are_ist_midnight_converted_to_utc():
    start, end = _day_bounds_utc(date(2026, 8, 18))

    # IST is UTC+5:30 -- 00:00 IST on the 18th is 18:30 UTC on the 17th.
    assert start == datetime(2026, 8, 17, 18, 30, tzinfo=UTC)
    assert end == datetime(2026, 8, 18, 18, 30, tzinfo=UTC)


# -- _sanitize_sheet_name -------------------------------------------------


def test_sanitize_sheet_name_replaces_illegal_excel_characters():
    assert _sanitize_sheet_name("NIFTY: 18/08/2026") == "NIFTY- 18-08-2026"


def test_sanitize_sheet_name_truncates_to_31_chars():
    assert len(_sanitize_sheet_name("X" * 50)) == 31


# -- export_trade_log_for_workspace ---------------------------------------


def test_zero_rows_returns_none_and_touches_no_file(tmp_path):
    result = export_trade_log_for_workspace(uuid.uuid4(), [], date(2026, 8, 18))

    assert result is None
    assert list(tmp_path.iterdir()) == []


def test_creates_workbook_with_correct_sheet_and_headers(tmp_path):
    workspace_id = uuid.uuid4()
    row = _row(workspace_id=workspace_id)

    path = export_trade_log_for_workspace(workspace_id, [row], date(2026, 8, 18))

    assert path == tmp_path / f"trade_log_{workspace_id}.xlsx"
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == ["orb"]  # sheet routes by strategy, not underlying/expiry
    ws = wb["orb"]
    assert ws.cell(row=1, column=1).value == "Strategy"
    assert ws.cell(row=2, column=1).value == "orb"
    assert ws.cell(row=1, column=2).value == "Underlying"
    assert ws.cell(row=2, column=2).value == "NIFTY"
    assert ws.cell(row=1, column=3).value == "Expiry"
    assert ws.cell(row=2, column=3).value == "2026-08-18"
    assert ws.cell(row=1, column=5).value == "Paper/Live"
    assert ws.cell(row=2, column=5).value == "paper"
    assert ws.cell(row=1, column=18).value == "VIX (at entry)"
    assert ws.cell(row=2, column=18).value == row.vix
    assert ws.cell(row=1, column=19).value == "OI (at entry)"
    assert ws.cell(row=2, column=19).value == row.oi
    assert ws.cell(row=1, column=20).value == "PCR - OI (at entry)"
    assert ws.cell(row=2, column=20).value == row.pcr_oi
    assert ws.cell(row=1, column=21).value == "PCR - Volume (at entry)"
    assert ws.cell(row=2, column=21).value == row.pcr_vol
    assert ws.cell(row=1, column=22).value == "Leg"
    assert ws.cell(row=2, column=22).value == "1/1"
    assert ws.cell(row=1, column=23).value == "Leg Kind"
    assert ws.cell(row=2, column=23).value == "—"
    assert ws.cell(row=1, column=24).value == "Trade ID (internal)"
    assert ws.cell(row=2, column=24).value == str(row.trade_outcome_id)


def test_none_env_metrics_write_as_blank_cells(tmp_path):
    """A trade that predates the VIX/PCR pipeline (or fired before either
    ever landed a value) must not crash the export -- None values render as
    blank cells, not a formatting error."""
    workspace_id = uuid.uuid4()
    row = _row(workspace_id=workspace_id, vix=None, oi=None, pcr_oi=None, pcr_vol=None)

    path = export_trade_log_for_workspace(workspace_id, [row], date(2026, 8, 18))

    ws = openpyxl.load_workbook(path)["orb"]
    assert ws.cell(row=2, column=18).value is None
    assert ws.cell(row=2, column=19).value is None
    assert ws.cell(row=2, column=20).value is None
    assert ws.cell(row=2, column=21).value is None


def test_routes_by_strategy_not_underlying_or_expiry(tmp_path):
    """A strategy's trades across different underlyings/expiries/cycles all
    land in its own single tab -- the actual point of the 2026-08-19 change
    (evaluate a strategy's own performance in one place)."""
    workspace_id = uuid.uuid4()
    this_week = _row(workspace_id=workspace_id, strategy="orb", expiry=date(2026, 8, 18))
    next_week = _row(workspace_id=workspace_id, strategy="orb", expiry=date(2026, 8, 25))
    other_underlying = _row(
        workspace_id=workspace_id, strategy="orb", underlying="BANKNIFTY", expiry=date(2026, 8, 25)
    )
    other_strategy = _row(workspace_id=workspace_id, strategy="vwap_pullback")

    path = export_trade_log_for_workspace(
        workspace_id,
        [this_week, next_week, other_underlying, other_strategy],
        date(2026, 8, 18),
    )

    wb = openpyxl.load_workbook(path)
    assert set(wb.sheetnames) == {"orb", "vwap_pullback"}
    assert wb["orb"].max_row == 4  # header + 3 distinct trades, one tab regardless of cycle
    assert wb["vwap_pullback"].max_row == 2  # header + 1 row


def test_paper_and_live_trades_share_the_same_strategy_tab(tmp_path):
    """Per explicit design decision: one tab per strategy with a Paper/Live
    column, not separate paper/live tabs -- lets a strategy's paper vs live
    behavior be compared side by side."""
    workspace_id = uuid.uuid4()
    paper_trade = _row(workspace_id=workspace_id, strategy="orb", trade_mode="paper")
    live_trade = _row(workspace_id=workspace_id, strategy="orb", trade_mode="live")

    path = export_trade_log_for_workspace(
        workspace_id, [paper_trade, live_trade], date(2026, 8, 18)
    )

    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == ["orb"]
    ws = wb["orb"]
    assert ws.cell(row=2, column=5).value == "paper"
    assert ws.cell(row=3, column=5).value == "live"


def test_reexporting_the_same_trades_does_not_duplicate_rows(tmp_path):
    workspace_id = uuid.uuid4()
    row = _row(workspace_id=workspace_id)

    export_trade_log_for_workspace(workspace_id, [row], date(2026, 8, 18))
    path = export_trade_log_for_workspace(workspace_id, [row], date(2026, 8, 18))

    wb = openpyxl.load_workbook(path)
    assert wb["orb"].max_row == 2  # still just header + 1 row


def test_new_trades_append_alongside_already_exported_ones(tmp_path):
    workspace_id = uuid.uuid4()
    first = _row(workspace_id=workspace_id)
    second = _row(workspace_id=workspace_id)

    export_trade_log_for_workspace(workspace_id, [first], date(2026, 8, 18))
    path = export_trade_log_for_workspace(workspace_id, [first, second], date(2026, 8, 18))

    wb = openpyxl.load_workbook(path)
    assert wb["orb"].max_row == 3  # header + 2 distinct trades


# The exact 22-column header this file used before the multi-leg exit
# engine's "Leg"/"Leg Kind" columns were added -- a real, already-exporting
# `trade_log_<workspace_id>.xlsx` on disk looks exactly like this. Pinned as
# a literal snapshot (not derived from the current `_HEADERS`) so this test
# keeps testing the real historical regression even if `_HEADERS` changes
# again later.
_OLD_22_COLUMN_HEADERS = [
    "Strategy", "Underlying", "Expiry", "Execution Mode", "Paper/Live",
    "Contract Symbol", "Option Type", "Strike", "Side", "Qty",
    "Entry Price", "Entry Time (IST)", "Exit Price", "Exit Time (IST)",
    "Exit Reason", "Realized PnL", "Slippage", "VIX (at entry)",
    "OI (at entry)", "PCR - OI (at entry)", "PCR - Volume (at entry)",
    "Trade ID (internal)",
]


def test_appending_to_a_pre_existing_old_schema_sheet_stays_backward_compatible(tmp_path):
    """A real trade_log_<workspace_id>.xlsx already exists in production with
    the old 22-column header (no Leg/Leg Kind). Appending to it must not
    shift columns, must not add Leg/Leg Kind to that sheet, must not
    duplicate the pre-existing row, and must still recognize a freshly
    appended row as already-exported on a second run -- see exporter.py's
    own module comment on `_HEADERS` for the incident this guards against."""
    workspace_id = uuid.uuid4()
    path = tmp_path / f"trade_log_{workspace_id}.xlsx"

    old_trade_id = uuid.uuid4()
    wb = openpyxl.Workbook()
    del wb["Sheet"]
    ws = wb.create_sheet(title="orb")
    ws.append(_OLD_22_COLUMN_HEADERS)
    ws.append(
        [
            "orb", "NIFTY", "2026-08-17", "auto", "paper",
            "NIFTY17AUG26C24000", "CE", 24000.0, "buy", 25,
            80.0, datetime(2026, 8, 17, 5, 0), 100.0, datetime(2026, 8, 17, 6, 0),
            "target", 500.0, 1.5, 13.5, 125000, 0.9, 1.1,
            str(old_trade_id),
        ]
    )
    wb.save(path)

    new_row = _row(workspace_id=workspace_id, trade_id=uuid.uuid4())
    export_trade_log_for_workspace(workspace_id, [new_row], date(2026, 8, 18))

    wb2 = openpyxl.load_workbook(path)
    ws2 = wb2["orb"]
    assert ws2.max_column == 22  # unchanged -- no Leg/Leg Kind added to this sheet
    assert ws2.max_row == 3  # header + old row + newly-appended row, no shift
    # The pre-existing row is untouched.
    assert ws2.cell(row=2, column=10).value == 25  # Qty
    assert ws2.cell(row=2, column=22).value == str(old_trade_id)  # Trade ID
    # The newly-appended row lands under the sheet's own (old) headers.
    assert ws2.cell(row=3, column=10).value == new_row.qty  # Qty
    assert ws2.cell(row=3, column=22).value == str(new_row.trade_outcome_id)  # Trade ID

    # Re-exporting both the old and new trade again must not duplicate
    # either — idempotency must survive on an old-schema sheet.
    export_trade_log_for_workspace(workspace_id, [new_row], date(2026, 8, 18))
    wb3 = openpyxl.load_workbook(path)
    assert wb3["orb"].max_row == 3


def test_separate_workspaces_get_separate_files(tmp_path):
    ws_a, ws_b = uuid.uuid4(), uuid.uuid4()

    path_a = export_trade_log_for_workspace(ws_a, [_row(workspace_id=ws_a)], date(2026, 8, 18))
    path_b = export_trade_log_for_workspace(ws_b, [_row(workspace_id=ws_b)], date(2026, 8, 18))

    assert path_a != path_b
    assert path_a.exists() and path_b.exists()


def test_permission_error_falls_back_to_csv(tmp_path, monkeypatch):
    def _raise_permission_error(self, path):
        raise PermissionError("file is open in Excel")

    monkeypatch.setattr(openpyxl.Workbook, "save", _raise_permission_error)

    workspace_id = uuid.uuid4()
    row = _row(workspace_id=workspace_id)

    result_path = export_trade_log_for_workspace(workspace_id, [row], date(2026, 8, 18))

    assert result_path is not None
    assert result_path.suffix == ".csv"
    assert result_path.exists()
    # The real .xlsx must never have been left in a half-written state.
    assert not (tmp_path / f"trade_log_{workspace_id}.xlsx").exists()

    with result_path.open(newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    assert rows[0] == _HEADERS
    assert rows[1][0] == "orb"  # Strategy
    assert rows[1][1] == "NIFTY"  # Underlying
    assert rows[1][4] == "paper"  # Paper/Live
