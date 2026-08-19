"""Ops-Hardening Phase 3: app.modules.reporting.exporter.fetch_completed_trades_
for_day's real DB-query correctness (day-boundary filtering, join
correctness) -- the pure Excel-write logic is covered separately in
tests/unit/test_trade_log_exporter.py against synthetic rows.
"""

from __future__ import annotations

import uuid
from datetime import UTC, date, datetime, timedelta

import pytest
from sqlalchemy.orm import Session

from app.domain.execution.models import (
    ExitReason,
    Order,
    OrderMode,
    OrderSide,
    OrderStatus,
    OrderType,
    Position,
    PositionStatus,
    TradeOutcome,
)
from app.domain.identity.models import BrokerAccount, BrokerAccountStatus, BrokerType, User
from app.domain.market.models import (
    Instrument,
    OptionChainSnapshot,
    OptionContract,
    OptionType,
    QuoteTick,
)
from app.domain.session.models import FundingMode, SafeMode, TradingSession
from app.domain.strategy.models import (
    ExecutionMode,
    Signal,
    SignalSide,
    StrategyConfig,
    StrategyRun,
    StrategyRunStatus,
    TradeIntent,
    TradeIntentStatus,
)
from app.modules.reporting.exporter import (
    export_completed_trades_for_day,
    fetch_completed_trades_for_day,
)

EXPIRY = date(2026, 8, 18)


@pytest.fixture
def broker_account(db: Session, workspace) -> BrokerAccount:
    account = BrokerAccount(
        id=uuid.uuid4(),
        workspace_id=workspace.id,
        broker_type=BrokerType.SHOONYA,
        label="export-query-test-account",
        credentials_ref="config/credentials/shoonya.env",
        status=BrokerAccountStatus.ACTIVE,
    )
    db.add(account)
    db.flush()
    return account


@pytest.fixture
def trading_session(db: Session, workspace, broker_account, user: User) -> TradingSession:
    ts = TradingSession(
        id=uuid.uuid4(),
        workspace_id=workspace.id,
        broker_account_id=broker_account.id,
        started_by_user_id=user.id,
        mode=SafeMode.PAPER_ONLY,
        started_at=datetime.now(UTC),
        budget_amount=1_000_000,
        daily_target_profit=1_000_000,
        daily_loss_cap=1_000_000,
        funding_mode=FundingMode.CASH,
    )
    db.add(ts)
    db.flush()
    return ts


@pytest.fixture
def instrument(db: Session) -> Instrument:
    inst = Instrument(id=uuid.uuid4(), symbol="NIFTY", exchange="NFO", lot_size=25, tick_size=0.05)
    db.add(inst)
    db.flush()
    return inst


@pytest.fixture
def option_contract(db: Session, instrument: Instrument) -> OptionContract:
    contract = OptionContract(
        id=uuid.uuid4(),
        instrument_id=instrument.id,
        expiry_date=EXPIRY,
        strike=24000,
        option_type=OptionType.CE,
        symbol="NIFTY18AUG26C24000",
    )
    db.add(contract)
    db.flush()
    return contract


def _seed_completed_trade(
    db: Session,
    *,
    workspace,
    user: User,
    trading_session: TradingSession,
    option_contract: OptionContract,
    strategy_name: str,
    closed_at: datetime,
    realized_pnl: float = 500.0,
    order_mode: OrderMode = OrderMode.PAPER,
) -> TradeOutcome:
    """Builds a full Signal -> TradeIntent -> Order(open) -> Position ->
    Order(close) -> TradeOutcome chain, bypassing the real execution engine
    since only the *query* joins are under test here. Respects the circular
    orders<->positions FK discipline this project's own CLAUDE.md documents
    (an opening Order carries trade_intent_id, a closing Order carries
    position_id, never both — ck_order_exactly_one_of_intent_or_position).
    """
    config = StrategyConfig(id=uuid.uuid4(), workspace_id=workspace.id, name=strategy_name)
    db.add(config)
    db.flush()

    run = StrategyRun(
        id=uuid.uuid4(),
        strategy_config_id=config.id,
        trading_session_id=trading_session.id,
        execution_mode=ExecutionMode.AUTO,
        status=StrategyRunStatus.SCANNING,
        started_at=closed_at,
        started_by_user_id=user.id,
    )
    db.add(run)
    db.flush()

    signal = Signal(
        id=uuid.uuid4(),
        workspace_id=workspace.id,
        strategy_config_id=config.id,
        strategy_run_id=run.id,
        trading_session_id=trading_session.id,
        option_contract_id=option_contract.id,
        side=SignalSide.BUY,
        entry_price=80.0,
        stop_price=60.0,
        target_price=120.0,
        qty_lots=1,
        generated_at=closed_at,
    )
    db.add(signal)
    db.flush()

    intent = TradeIntent(
        id=uuid.uuid4(),
        workspace_id=workspace.id,
        signal_id=signal.id,
        strategy_run_id=run.id,
        trading_session_id=trading_session.id,
        option_contract_id=option_contract.id,
        idempotency_key=f"test-{uuid.uuid4()}",
        side=SignalSide.BUY,
        qty_lots=1,
        entry_price=80.0,
        stop_price=60.0,
        target_price=120.0,
        status=TradeIntentStatus.DISPATCHED,
        created_at=closed_at,
        dispatched_at=closed_at,
    )
    db.add(intent)
    db.flush()

    opening_order = Order(
        id=uuid.uuid4(),
        workspace_id=workspace.id,
        trading_session_id=trading_session.id,
        option_contract_id=option_contract.id,
        trade_intent_id=intent.id,
        idempotency_key=f"test-open-{uuid.uuid4()}",
        mode=order_mode,
        side=OrderSide.BUY,
        order_type=OrderType.MARKET,
        qty=25,
        status=OrderStatus.FILLED,
        filled_qty=25,
        avg_fill_price=80.0,
        submitted_at=closed_at,
        updated_at=closed_at,
    )
    db.add(opening_order)
    db.flush()

    position = Position(
        id=uuid.uuid4(),
        workspace_id=workspace.id,
        trading_session_id=trading_session.id,
        option_contract_id=option_contract.id,
        trade_intent_id=intent.id,
        opening_order_id=opening_order.id,
        side=OrderSide.BUY,
        qty=25,
        entry_price=80.0,
        status=PositionStatus.CLOSED,
        opened_at=closed_at,
        closed_at=closed_at,
    )
    db.add(position)
    db.flush()

    closing_order = Order(
        id=uuid.uuid4(),
        workspace_id=workspace.id,
        trading_session_id=trading_session.id,
        option_contract_id=option_contract.id,
        position_id=position.id,
        idempotency_key=f"test-close-{uuid.uuid4()}",
        mode=OrderMode.PAPER,
        side=OrderSide.SELL,
        order_type=OrderType.MARKET,
        qty=25,
        status=OrderStatus.FILLED,
        filled_qty=25,
        avg_fill_price=100.0,
        submitted_at=closed_at,
        updated_at=closed_at,
    )
    db.add(closing_order)
    db.flush()
    position.closing_order_id = closing_order.id
    db.flush()

    outcome = TradeOutcome(
        id=uuid.uuid4(),
        workspace_id=workspace.id,
        trading_session_id=trading_session.id,
        position_id=position.id,
        trade_intent_id=intent.id,
        entry_price=80.0,
        exit_price=100.0,
        qty=25,
        realized_pnl=realized_pnl,
        slippage=1.5,
        exit_reason=ExitReason.TARGET,
        closed_at=closed_at,
    )
    db.add(outcome)
    db.flush()
    return outcome


def test_fetches_only_trades_closed_within_the_ist_day(
    db: Session, workspace, user, trading_session, option_contract
):
    # 2026-08-18 00:00 IST == 2026-08-17 18:30 UTC. One trade just inside the
    # day, one just before it, one just after it.
    inside = _seed_completed_trade(
        db,
        workspace=workspace,
        user=user,
        trading_session=trading_session,
        option_contract=option_contract,
        strategy_name="orb-inside",
        closed_at=datetime(2026, 8, 17, 18, 30, tzinfo=UTC),
    )
    _seed_completed_trade(
        db,
        workspace=workspace,
        user=user,
        trading_session=trading_session,
        option_contract=option_contract,
        strategy_name="orb-before",
        closed_at=datetime(2026, 8, 17, 18, 29, 59, tzinfo=UTC),
    )
    _seed_completed_trade(
        db,
        workspace=workspace,
        user=user,
        trading_session=trading_session,
        option_contract=option_contract,
        strategy_name="orb-after",
        closed_at=datetime(2026, 8, 18, 18, 30, tzinfo=UTC),
    )

    rows = fetch_completed_trades_for_day(db, date(2026, 8, 18))

    assert {r.trade_outcome_id for r in rows} == {inside.id}


def test_row_carries_strategy_execution_mode_and_cycle_fields(
    db: Session, workspace, user, trading_session, option_contract
):
    _seed_completed_trade(
        db,
        workspace=workspace,
        user=user,
        trading_session=trading_session,
        option_contract=option_contract,
        strategy_name="vwap_pullback",
        closed_at=datetime(2026, 8, 18, 6, 0, tzinfo=UTC),
        realized_pnl=750.0,
    )

    rows = fetch_completed_trades_for_day(db, date(2026, 8, 18))

    assert len(rows) == 1
    row = rows[0]
    assert row.strategy_name == "vwap_pullback"
    assert row.execution_mode == "auto"
    assert row.trade_mode == "paper"
    assert row.underlying_symbol == "NIFTY"
    assert row.expiry_date == EXPIRY
    assert row.workspace_id == workspace.id
    assert row.realized_pnl == 750.0
    assert row.entry_price == 80.0
    assert row.exit_price == 100.0


def test_row_carries_env_metrics_as_of_the_trades_own_entry_time(
    db: Session, workspace, user, trading_session, instrument, option_contract
):
    """2026-08-19: VIX/PCR must be reconstructed as of `position.opened_at`
    (the trade's own entry time), not whatever is current when the export
    runs -- a later VIX tick / option-chain snapshot must not leak in.
    """
    entry_time = datetime(2026, 8, 18, 6, 0, tzinfo=UTC)
    vix_instrument = Instrument(
        id=uuid.uuid4(), symbol="INDIA VIX", exchange="NSE", lot_size=1, tick_size=0.05
    )
    db.add(vix_instrument)
    db.flush()

    db.add_all(
        [
            QuoteTick(
                id=uuid.uuid4(),
                instrument_id=vix_instrument.id,
                ltp=13.0,
                bid=13.0,
                ask=13.0,
                volume=0,
                ts=entry_time - timedelta(minutes=1),
            ),
            QuoteTick(
                id=uuid.uuid4(),
                instrument_id=vix_instrument.id,
                ltp=20.0,  # after entry -- must not be picked up
                bid=20.0,
                ask=20.0,
                volume=0,
                ts=entry_time + timedelta(minutes=10),
            ),
            OptionChainSnapshot(
                id=uuid.uuid4(),
                instrument_id=instrument.id,
                expiry_date=EXPIRY,
                ts=entry_time - timedelta(seconds=30),
                chain_data=[
                    {
                        "contract_symbol": "NIFTY18AUG26C24000",
                        "option_type": "CE",
                        "oi": 1000,
                        "volume": 100,
                    },
                    {"option_type": "PE", "oi": 500, "volume": 50},
                ],
            ),
            OptionChainSnapshot(
                id=uuid.uuid4(),
                instrument_id=instrument.id,
                expiry_date=EXPIRY,
                ts=entry_time + timedelta(minutes=10),  # after entry -- must not be picked up
                chain_data=[
                    {
                        "contract_symbol": "NIFTY18AUG26C24000",
                        "option_type": "CE",
                        "oi": 9999,  # after entry -- must not be picked up
                        "volume": 100,
                    },
                    {"option_type": "PE", "oi": 1000, "volume": 100},
                ],
            ),
        ]
    )
    db.flush()

    _seed_completed_trade(
        db,
        workspace=workspace,
        user=user,
        trading_session=trading_session,
        option_contract=option_contract,
        strategy_name="orb",
        closed_at=entry_time,
    )

    rows = fetch_completed_trades_for_day(db, date(2026, 8, 18))

    assert len(rows) == 1
    row = rows[0]
    assert row.vix == 13.0
    assert row.oi == 1000  # the traded contract's own OI as of entry, not the later 9999
    assert row.pcr_oi == 0.5
    assert row.pcr_vol == 0.5


def test_row_env_metrics_are_none_when_nothing_was_known_yet(
    db: Session, workspace, user, trading_session, option_contract
):
    _seed_completed_trade(
        db,
        workspace=workspace,
        user=user,
        trading_session=trading_session,
        option_contract=option_contract,
        strategy_name="orb",
        closed_at=datetime(2026, 8, 18, 6, 0, tzinfo=UTC),
    )

    rows = fetch_completed_trades_for_day(db, date(2026, 8, 18))

    assert len(rows) == 1
    assert rows[0].vix is None
    assert rows[0].oi is None
    assert rows[0].pcr_oi is None
    assert rows[0].pcr_vol is None


def test_row_carries_the_opening_orders_own_paper_or_live_mode(
    db: Session, workspace, user, trading_session, option_contract
):
    """Order.mode on the position's opening order -- distinct from
    execution_mode (auto/manual approval), which is orthogonal to this."""
    _seed_completed_trade(
        db,
        workspace=workspace,
        user=user,
        trading_session=trading_session,
        option_contract=option_contract,
        strategy_name="orb-paper",
        closed_at=datetime(2026, 8, 18, 6, 0, tzinfo=UTC),
        order_mode=OrderMode.PAPER,
    )
    _seed_completed_trade(
        db,
        workspace=workspace,
        user=user,
        trading_session=trading_session,
        option_contract=option_contract,
        strategy_name="orb-live",
        closed_at=datetime(2026, 8, 18, 7, 0, tzinfo=UTC),
        order_mode=OrderMode.LIVE,
    )

    rows = fetch_completed_trades_for_day(db, date(2026, 8, 18))

    by_strategy = {r.strategy_name: r.trade_mode for r in rows}
    assert by_strategy == {"orb-paper": "paper", "orb-live": "live"}


def test_export_completed_trades_for_day_writes_a_workbook(
    db: Session, workspace, user, trading_session, option_contract, tmp_path, monkeypatch
):
    from contextlib import contextmanager

    from app.modules.reporting import exporter as exporter_module

    monkeypatch.setattr(exporter_module, "REPORTS_DIR", tmp_path)

    _seed_completed_trade(
        db,
        workspace=workspace,
        user=user,
        trading_session=trading_session,
        option_contract=option_contract,
        strategy_name="orb",
        closed_at=datetime(2026, 8, 18, 6, 0, tzinfo=UTC),
    )

    @contextmanager
    def _same_session():
        yield db

    export_completed_trades_for_day(date(2026, 8, 18), session_factory=_same_session)

    import openpyxl

    path = tmp_path / f"trade_log_{workspace.id}.xlsx"
    assert path.exists()
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == ["orb"]
    assert wb["orb"].cell(row=2, column=1).value == "orb"


def test_export_completed_trades_for_day_no_trades_touches_no_file(
    db: Session, tmp_path, monkeypatch
):
    from contextlib import contextmanager

    from app.modules.reporting import exporter as exporter_module

    monkeypatch.setattr(exporter_module, "REPORTS_DIR", tmp_path)

    @contextmanager
    def _same_session():
        yield db

    export_completed_trades_for_day(date(2026, 8, 18), session_factory=_same_session)

    assert list(tmp_path.iterdir()) == []
